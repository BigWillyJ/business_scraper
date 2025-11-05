"""
Small Local Business Scraper with Zip Code Search
Uses DeepSeek R1 14B locally to find and extract small independent service businesses
"""

import requests
from bs4 import BeautifulSoup
import json
import csv
import time
from typing import List, Dict, Optional
import re
from urllib.parse import quote_plus, urljoin, urlparse
import ollama
from datetime import datetime


class SmallBusinessScraper:
    def __init__(self, model_name: str = "deepseek-r1:14b"):
        """
        Initialize scraper with local DeepSeek model

        Args:
            model_name: Ollama model name (default: deepseek-r1:14b)
        """
        self.model_name = model_name
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        # Service-based business types we're targeting
        self.service_types = [
            "plumbers", "plumbing services",
            "hvac companies", "heating and cooling",
            "beauty salons", "hair salons",
            "massage therapists", "spa services",
            "furniture stores",
            "electricians", "electrical services",
            "carpenters", "carpentry services",
            "landscaping companies", "lawn care",
            "house cleaning services", "maid services",
            "pest control services",
            "locksmiths",
            "roofing companies", "roofers",
            "painting contractors",
            "flooring companies",
            "appliance repair"
        ]

        print(f"✓ Initialized with model: {model_name}")
        self._test_model_connection()

    def _test_model_connection(self):
        """Test connection to local Ollama model"""
        try:
            response = ollama.chat(
                model=self.model_name,
                messages=[{'role': 'user', 'content': 'Reply with just: OK'}]
            )
            print(f"✓ Model connection successful")
        except Exception as e:
            print(f"✗ Error connecting to model: {e}")
            print(f"  Make sure Ollama is running and model is installed:")
            print(f"  ollama run {self.model_name}")
            raise

    def fetch_business_urls_by_zipcode(self, zipcode: str, max_businesses: int = 20) -> List[str]:
        """
        Fetch business URLs for service businesses in a given zip code

        Args:
            zipcode: 5-digit US zip code
            max_businesses: Maximum number of business URLs to fetch

        Returns:
            List of business website URLs
        """
        all_urls = set()

        print(f"\n{'='*60}")
        print(f"🔍 Searching for service businesses in ZIP code: {zipcode}")
        print(f"{'='*60}\n")

        # Search for multiple business types
        search_count = 0
        for service_type in self.service_types:
            if len(all_urls) >= max_businesses:
                break

            search_count += 1
            print(f"[{search_count}/{len(self.service_types)}] Searching: {service_type} near {zipcode}")

            urls = self._search_google_oxylabs(f"{service_type} near {zipcode}")

            for url in urls:
                if len(all_urls) >= max_businesses:
                    break

                # Filter out non-business sites
                if self._is_valid_business_url(url):
                    all_urls.add(url)
                    print(f"  ✓ Found: {url}")

            # Rate limiting for searches
            time.sleep(2)

        url_list = list(all_urls)
        print(f"\n✓ Total unique business URLs found: {len(url_list)}")
        return url_list

    def _search_google_oxylabs(self, query: str, num_results: int = 10) -> List[str]:
        """
        Search Google using Oxylabs SERP Scraper API

        Setup:
        1. Sign up at https://oxylabs.io
        2. Get your username and password from dashboard
        3. Add credentials to oxylabs_config below
        """

        # ===== OXYLABS CONFIGURATION =====
        oxylabs_config = {
            'username': 'filljoey_scrapes_RDkMh',
            'password': 'iosdf92~Icks',
        }

        # Check if credentials are configured
        if oxylabs_config['username'] == 'YOUR_OXYLABS_USERNAME':
            print(f"  ⚠ Oxylabs credentials not configured!")
            print(f"  → Add your Oxylabs username and password in _search_google_oxylabs()")
            return []

        try:
            # Oxylabs SERP API endpoint
            url = 'https://realtime.oxylabs.io/v1/queries'

            # Request payload for Google Search
            payload = {
                'source': 'google_search',
                'query': query,
                'parse': True,
                'geo_location': 'United States',
                'pages': 1,
            }

            # Make request with Basic Auth
            response = requests.post(
                url,
                json=payload,
                auth=(oxylabs_config['username'], oxylabs_config['password']),
                timeout=30
            )

            if response.status_code == 401:
                print(f"  ✗ Oxylabs authentication failed. Check your credentials.")
                return []

            response.raise_for_status()
            data = response.json()

            # Extract organic search results
            urls = []
            results = data.get('results', [])

            for result in results:
                content = result.get('content', {})
                organic_results = content.get('results', {}).get('organic', [])

                for item in organic_results:
                    url = item.get('url')
                    if url and self._is_valid_business_url(url):
                        urls.append(url)
                        if len(urls) >= num_results:
                            break

                if len(urls) >= num_results:
                    break

            return urls[:num_results]

        except requests.exceptions.RequestException as e:
            print(f"  ✗ Oxylabs API error: {e}")
            return []
        except Exception as e:
            print(f"  ⚠ Unexpected error: {e}")
            return []

    def _is_valid_business_url(self, url: str) -> bool:
        """Check if URL is likely a legitimate business website"""
        try:
            parsed = urlparse(url)
            domain = parsed.netloc.lower()

            # Exclude common non-business sites
            excluded_domains = [
                'google.com', 'facebook.com', 'yelp.com', 'yellowpages.com',
                'bbb.org', 'angieslist.com', 'thumbtack.com', 'homeadvisor.com',
                'linkedin.com', 'instagram.com', 'twitter.com', 'youtube.com',
                'wikipedia.org', 'bing.com', 'yahoo.com', 'mapquest.com',
                'tripadvisor.com', 'foursquare.com', 'nextdoor.com'
            ]

            for excluded in excluded_domains:
                if excluded in domain:
                    return False

            # Must have a proper domain
            if '.' not in domain:
                return False

            return True

        except:
            return False

    def is_small_independent_business(self, business_data: Dict) -> bool:
        """
        Use DeepSeek to determine if business is small, independent, and service-based
        Excludes chains, franchises, manufacturers, and retail stores
        """
        prompt = f"""Analyze this business and determine if it meets ALL these criteria:

MUST BE:
- Small, independent business (not a chain or franchise)
- Service-based (provides services, not just selling products)
- Local trades like: plumbing, HVAC, beauty salons, massage therapists, furniture stores, electricians, carpenters, landscaping, cleaning, pest control, locksmiths, roofing, painting, etc.

MUST NOT BE:
- Part of a large chain or national franchise
- Manufacturer or factory
- Retail store (unless it's a local furniture store with services)
- Restaurant or food service
- Medical/dental office
- Real estate agency
- Bank or financial institution
- Directory or listing site

Business Information:
Name: {business_data.get('business_name', 'Unknown')}
Description: {business_data.get('description', 'N/A')}
Website: {business_data.get('website', 'N/A')}

Respond with ONLY a JSON object:
{{
    "is_small_independent": true/false,
    "is_service_based": true/false,
    "is_chain_or_franchise": true/false,
    "business_type": "brief category like 'plumbing service' or 'beauty salon'",
    "reasoning": "brief explanation"
}}"""

        try:
            response = ollama.chat(
                model=self.model_name,
                messages=[{'role': 'user', 'content': prompt}]
            )

            result_text = response['message']['content']

            # Extract JSON from response
            json_match = re.search(r'\{.*\}', result_text, re.DOTALL)
            if json_match:
                analysis = json.loads(json_match.group())

                # Must be small independent AND service-based AND NOT a chain
                is_qualified = (
                    analysis.get('is_small_independent', False) and
                    analysis.get('is_service_based', False) and
                    not analysis.get('is_chain_or_franchise', True)
                )

                if is_qualified:
                    print(f"  ✓ Qualified: {analysis.get('business_type', 'service business')}")
                else:
                    print(f"  ✗ Filtered out: {analysis.get('reasoning', 'Does not meet criteria')}")

                return is_qualified
            else:
                return False

        except Exception as e:
            print(f"  ⚠ Error analyzing business: {e}")
            return False

    def scrape_business_page(self, url: str) -> Optional[str]:
        """Scrape a business webpage and return raw HTML content"""
        try:
            response = requests.get(url, headers=self.headers, timeout=15)
            response.raise_for_status()
            return response.text
        except Exception as e:
            print(f"  ✗ Error scraping: {e}")
            return None

    def extract_business_info(self, html_content: str, url: str) -> Dict:
        """Use DeepSeek to extract structured business information from HTML"""
        soup = BeautifulSoup(html_content, 'html.parser')

        # Extract emails - prioritize mailto: links first (most reliable)
        mailto_pattern = r'mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
        mailto_emails = re.findall(mailto_pattern, html_content, re.IGNORECASE)

        # Then look for plain text emails
        email_pattern = r'\b([a-zA-Z0-9][a-zA-Z0-9._%+-]*@[a-zA-Z0-9][a-zA-Z0-9.-]*\.[a-zA-Z]{2,})\b'
        text_emails = re.findall(email_pattern, soup.get_text())

        # Combine and filter out invalid ones
        all_emails = mailto_emails + text_emails
        valid_emails = []

        for email in all_emails:
            email = email.lower().strip()

            # Skip if it contains file extensions or image patterns
            if re.search(r'\.(png|jpg|jpeg|gif|svg|webp|pdf|js|css|woff|ttf|ico)', email):
                continue
            if '@2x' in email or '@3x' in email:
                continue

            # Must have valid domain
            if '@' not in email:
                continue

            parts = email.split('@')
            if len(parts) != 2:
                continue

            domain = parts[1]

            # Skip weird domains
            if len(domain) < 4 or '/' in domain or '\\' in domain:
                continue

            # Must end with common valid TLD
            valid_tlds = ['com', 'net', 'org', 'edu', 'gov', 'co', 'us', 'io', 'biz', 'info']
            if not any(domain.endswith('.' + tld) for tld in valid_tlds):
                continue

            valid_emails.append(email)

        unique_emails = list(set(valid_emails))

        # Separate owner vs business emails
        owner_emails = [e for e in unique_emails if not any(prefix in e for prefix in ['info@', 'contact@', 'sales@', 'support@', 'hello@', 'admin@', 'service@', 'office@'])]
        business_emails = [e for e in unique_emails if any(prefix in e for prefix in ['info@', 'contact@', 'sales@', 'support@', 'hello@', 'admin@', 'service@', 'office@'])]

        # Extract phone numbers from entire HTML
        phone_patterns = [
            r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\d{3}[-.\s]\d{3}[-.\s]\d{4}',
            r'\(\d{3}\)\s?\d{3}-\d{4}',
        ]

        phones = []
        for pattern in phone_patterns:
            phones.extend(re.findall(pattern, html_content))

        # Clean, validate and format phones
        cleaned_phones = []
        for phone in phones:
            # Remove all non-digit characters to validate
            digits_only = re.sub(r'\D', '', phone)

            # Must be exactly 10 digits for valid US phone
            if len(digits_only) == 10:
                # Format as (XXX) XXX-XXXX
                formatted = f"({digits_only[:3]}) {digits_only[3:6]}-{digits_only[6:]}"
                if formatted not in cleaned_phones:
                    cleaned_phones.append(formatted)

        main_phone = cleaned_phones[0] if cleaned_phones else None

        # Extract address patterns
        address_pattern = r'\d+\s+[A-Z][a-zA-Z\s]+(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|Drive|Dr|Court|Ct|Way)\.?'
        addresses = re.findall(address_pattern, soup.get_text())
        main_address = addresses[0] if addresses else None

        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()

        # Get clean text
        text = soup.get_text(separator='\n', strip=True)
        text = '\n'.join(line for line in text.split('\n') if line.strip())
        text = text[:12000]

        # Get meta description
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        description = meta_desc.get('content', '') if meta_desc else ''

        # Create contact hints for the AI
        contact_hints = f"""
PRE-EXTRACTED CONTACT INFO (USE THESE):
- Emails: {', '.join(unique_emails[:10]) if unique_emails else 'None'}
- Owner emails: {', '.join(owner_emails[:5]) if owner_emails else 'None'}  
- Business emails: {', '.join(business_emails[:5]) if business_emails else 'None'}
- Phones: {', '.join(cleaned_phones[:3]) if cleaned_phones else 'None'}
- Address: {main_address if main_address else 'None'}
"""

        prompt = f"""Extract business information from this webpage. Return ONLY valid JSON.

{contact_hints}

PRIORITY: Phone and email are MOST IMPORTANT. Check footer and header sections carefully.

Required fields (use null if not found):
- business_name: Company name
- owner_name: Full name of owner
- owner_email: Personal email with name (NOT info@/contact@)
- business_email: Generic business email (info@, contact@, etc.)
- address: Full street address
- city: City name
- state: 2-letter state code
- zip_code: ZIP code
- phone: Phone number (USE PRE-EXTRACTED IF AVAILABLE)
- website: {url}
- description: Brief description (2-3 sentences)
- services: Array of services

CRITICAL: If phones/emails are listed above, you MUST use them. Do NOT create fake contact info.

Webpage URL: {url}
Meta: {description}

Text:
{text}

Return ONLY JSON, no other text."""

        try:
            response = ollama.chat(
                model=self.model_name,
                messages=[{'role': 'user', 'content': prompt}]
            )

            result_text = response['message']['content']

            # Extract JSON
            json_match = re.search(r'\{.*\}', result_text, re.DOTALL)
            if json_match:
                business_data = json.loads(json_match.group())
                business_data['source_url'] = url

                # Fill in pre-extracted data if model missed it
                if not business_data.get('business_email') and business_emails:
                    business_data['business_email'] = business_emails[0]
                if not business_data.get('owner_email') and owner_emails:
                    business_data['owner_email'] = owner_emails[0]
                if not business_data.get('phone') and main_phone:
                    business_data['phone'] = main_phone
                if not business_data.get('address') and main_address:
                    business_data['address'] = main_address

                return business_data
            else:
                print(f"  ⚠ Could not parse JSON from model response")
                return {
                    'source_url': url,
                    'business_email': business_emails[0] if business_emails else None,
                    'owner_email': owner_emails[0] if owner_emails else None,
                    'phone': main_phone,
                    'address': main_address
                }

        except Exception as e:
            print(f"  ✗ Error extracting business info: {e}")
            return {
                'source_url': url,
                'business_email': business_emails[0] if business_emails else None,
                'owner_email': owner_emails[0] if owner_emails else None,
                'phone': main_phone,
                'address': main_address
            }

    def scrape_businesses(self, urls: List[str]) -> List[Dict]:
        """
        Main method to scrape and filter businesses

        Args:
            urls: List of business website URLs to scrape

        Returns:
            List of qualified small independent service businesses
        """
        qualified_businesses = []

        print(f"\n{'='*60}")
        print(f"Starting scrape of {len(urls)} businesses")
        print(f"{'='*60}\n")

        for idx, url in enumerate(urls, 1):
            print(f"\n[{idx}/{len(urls)}] Processing: {url}")

            # Scrape the page
            html = self.scrape_business_page(url)
            if not html:
                continue

            # Extract business information
            print(f"  Extracting business info...")
            business_data = self.extract_business_info(html, url)

            if not business_data.get('business_name'):
                print(f"  ✗ Could not extract business name, skipping")
                continue

            # Check if it's a qualified small independent service business
            print(f"  Analyzing if qualified small business...")
            if self.is_small_independent_business(business_data):
                qualified_businesses.append(business_data)
                print(f"  ✓ Added to results")

            # Rate limiting - be respectful
            time.sleep(3)

        return qualified_businesses

    def save_results(self, businesses: List[Dict], zipcode: str):
        """Save qualified businesses to multiple file formats"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_filename = f"businesses_{zipcode}_{timestamp}"

        # 1. Save as JSON
        json_filename = f"{base_filename}.json"
        output = {
            'zipcode': zipcode,
            'total_qualified': len(businesses),
            'scraped_at': time.strftime('%Y-%m-%d %H:%M:%S'),
            'businesses': businesses
        }

        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        # 2. Save as CSV
        csv_filename = f"{base_filename}.csv"
        if businesses:
            # Get all possible keys from all businesses
            all_keys = set()
            for biz in businesses:
                all_keys.update(biz.keys())

            # Convert services list to string for CSV
            csv_businesses = []
            for biz in businesses:
                biz_copy = biz.copy()
                if 'services' in biz_copy and isinstance(biz_copy['services'], list):
                    biz_copy['services'] = ', '.join(biz_copy['services'])
                csv_businesses.append(biz_copy)

            fieldnames = ['business_name', 'owner_name', 'owner_email', 'business_email',
                         'phone', 'address', 'city', 'state', 'zip_code', 'website',
                         'description', 'services', 'business_type', 'source_url']

            with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(csv_businesses)

        # 3. Try to save as Excel (optional - requires openpyxl)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            xlsx_filename = f"{base_filename}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Businesses {zipcode}"

            # Headers
            headers = ['Business Name', 'Owner Name', 'Owner Email', 'Business Email',
                      'Phone', 'Address', 'City', 'State', 'ZIP', 'Website',
                      'Description', 'Services', 'Type', 'Source URL']

            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Add data
            for row, biz in enumerate(businesses, 2):
                ws.cell(row=row, column=1, value=biz.get('business_name'))
                ws.cell(row=row, column=2, value=biz.get('owner_name'))
                ws.cell(row=row, column=3, value=biz.get('owner_email'))
                ws.cell(row=row, column=4, value=biz.get('business_email'))
                ws.cell(row=row, column=5, value=biz.get('phone'))
                ws.cell(row=row, column=6, value=biz.get('address'))
                ws.cell(row=row, column=7, value=biz.get('city'))
                ws.cell(row=row, column=8, value=biz.get('state'))
                ws.cell(row=row, column=9, value=biz.get('zip_code'))
                ws.cell(row=row, column=10, value=biz.get('website'))
                ws.cell(row=row, column=11, value=biz.get('description'))

                services = biz.get('services', [])
                if isinstance(services, list):
                    ws.cell(row=row, column=12, value=', '.join(services))
                else:
                    ws.cell(row=row, column=12, value=services)

                ws.cell(row=row, column=13, value=biz.get('business_type'))
                ws.cell(row=row, column=14, value=biz.get('source_url'))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(xlsx_filename)

            print(f"\n{'='*60}")
            print(f"✓ Saved {len(businesses)} businesses to:")
            print(f"  📄 {json_filename} (JSON format)")
            print(f"  📊 {csv_filename} (CSV format)")
            print(f"  📈 {xlsx_filename} (Excel format)")
            print(f"{'='*60}")

        except ImportError:
            print(f"\n{'='*60}")
            print(f"✓ Saved {len(businesses)} businesses to:")
            print(f"  📄 {json_filename} (JSON format)")
            print(f"  📊 {csv_filename} (CSV format)")
            print(f"\n💡 Tip: Install openpyxl for Excel format:")
            print(f"  pip install openpyxl")
            print(f"{'='*60}")

    def print_summary(self, businesses: List[Dict]):
        """Print a summary of scraped businesses"""
        print(f"\n{'='*60}")
        print(f"SUMMARY - Found {len(businesses)} Qualified Small Businesses")
        print(f"{'='*60}\n")

        for idx, biz in enumerate(businesses, 1):
            print(f"{idx}. {biz.get('business_name', 'Unknown')}")
            print(f"   Owner: {biz.get('owner_name', 'N/A')}")
            print(f"   Owner Email: {biz.get('owner_email', 'N/A')}")
            print(f"   Business Email: {biz.get('business_email', 'N/A')}")
            print(f"   Type: {biz.get('business_type', 'N/A')}")
            print(f"   Phone: {biz.get('phone', 'N/A')}")
            print(f"   Address: {biz.get('address', 'N/A')}, {biz.get('city', 'N/A')}, {biz.get('state', 'N/A')}")
            print(f"   Website: {biz.get('website', 'N/A')}")
            print()

    def run_by_zipcode(self, zipcode: str, max_businesses: int = 20):
        """
        Complete workflow: Search by zip code, scrape, filter, and save

        Args:
            zipcode: 5-digit US zip code
            max_businesses: Maximum number of businesses to find and scrape
        """
        # Validate zip code
        if not re.match(r'^\d{5}$', zipcode):
            print(f"✗ Invalid zip code: {zipcode}. Please use 5-digit format (e.g., '98052')")
            return

        # Step 1: Fetch business URLs
        urls = self.fetch_business_urls_by_zipcode(zipcode, max_businesses)

        if not urls:
            print("\n✗ No business URLs found. Try a different zip code or check your internet connection.")
            return

        # Step 2: Scrape and filter businesses
        qualified_businesses = self.scrape_businesses(urls)

        # Step 3: Save results
        if qualified_businesses:
            self.save_results(qualified_businesses, zipcode)
            self.print_summary(qualified_businesses)
        else:
            print("\n⚠ No qualified small independent service businesses found.")


def main():
    """Main entry point - Configure Oxylabs and enter zip code!"""

    # Initialize scraper with DeepSeek R1 14B
    print("Initializing Small Business Scraper with Oxylabs...")
    scraper = SmallBusinessScraper(model_name="deepseek-r1:14b")

    # ===== SETUP INSTRUCTIONS =====
    print("\n" + "="*60)
    print("OXYLABS SETUP:")
    print("="*60)
    print("1. Sign up at https://oxylabs.io")
    print("2. Get your username and password from the dashboard")
    print("3. Edit _search_google_oxylabs() method (around line 135)")
    print("4. Replace 'YOUR_OXYLABS_USERNAME' and 'YOUR_OXYLABS_PASSWORD'")
    print("5. Set your ZIP code below and run the script")
    print("="*60 + "\n")

    # ===== AUTO-SEARCH BY ZIP CODE =====
    ZIPCODE = "80003"  # Change to your target zip code
    MAX_BUSINESSES = 150  # How many businesses to find

    scraper.run_by_zipcode(ZIPCODE, max_businesses=MAX_BUSINESSES)


if __name__ == "__main__":
    main()